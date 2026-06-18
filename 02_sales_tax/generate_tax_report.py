import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

db_path = r'C:\Users\marsh\Downloads\Upwork\finance-analyst-portfolio\data\nexoria.db'
output_excel = r'C:\Users\marsh\Downloads\Upwork\finance-analyst-portfolio\02_sales_tax\sales_tax_report.xlsx'

# 1. Pull data from SQLite using the order-level query
conn = sqlite3.connect(db_path)
query = """
SELECT 
    o.order_id,
    o.order_date,
    o.customer_id,
    o.state,
    t.nexus AS has_nexus,
    o.taxable_amount,
    t.state_tax_rate AS correct_rate,
    ROUND(o.taxable_amount * t.state_tax_rate * t.nexus, 2) AS expected_tax,
    o.sales_tax AS collected_tax,
    ROUND(o.sales_tax - (o.taxable_amount * t.state_tax_rate * t.nexus), 2) AS discrepancy
FROM orders o
JOIN state_tax_rates t ON o.state = t.state_code;
"""
df = pd.read_sql_query(query, conn)
conn.close()

# 2. Setup Excel Workbook
wb = Workbook()
ws_dash = wb.active
ws_dash.title = "Tax Dashboard"
ws_data = wb.create_sheet(title="Raw Data")

# 3. Write Raw Data
for r in dataframe_to_rows(df, index=False, header=True):
    ws_data.append(r)

# Format Data Sheet Headers
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True)
for cell in ws_data[1]:
    cell.fill = header_fill
    cell.font = white_font

# 4. Build Dashboard Sheet
ws_dash.sheet_view.showGridLines = False

# Title
ws_dash['A1'] = "Nexoria Commerce - State Sales Tax Reconciliation"
ws_dash['A1'].font = Font(size=16, bold=True, color="1F4E78")

# Get unique states that have data
states = sorted(df['state'].unique())

# Dashboard Headers
headers = ["State", "Total Taxable Sales", "Expected Tax", "Collected Tax", "Discrepancy", "Status"]
ws_dash.append([])
ws_dash.append(headers)

for cell in ws_dash[3]:
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center")

# Define Styles
currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
border_style = Border(left=Side(style='thin', color='D9D9D9'),
                      right=Side(style='thin', color='D9D9D9'),
                      top=Side(style='thin', color='D9D9D9'),
                      bottom=Side(style='thin', color='D9D9D9'))

# Add state rows with Excel formulas instead of hardcoded aggregations!
row_num = 4
for state in states:
    ws_dash[f'A{row_num}'] = state
    # B: Total Taxable Sales -> sumifs column F where column D = state
    ws_dash[f'B{row_num}'] = f"=SUMIFS('Raw Data'!F:F, 'Raw Data'!D:D, A{row_num})"
    # C: Expected Tax -> sumifs column H where column D = state
    ws_dash[f'C{row_num}'] = f"=SUMIFS('Raw Data'!H:H, 'Raw Data'!D:D, A{row_num})"
    # D: Collected Tax -> sumifs column I where column D = state
    ws_dash[f'D{row_num}'] = f"=SUMIFS('Raw Data'!I:I, 'Raw Data'!D:D, A{row_num})"
    # E: Discrepancy -> sumifs column J where column D = state
    ws_dash[f'E{row_num}'] = f"=SUMIFS('Raw Data'!J:J, 'Raw Data'!D:D, A{row_num})"
    # F: Status -> IF discrepancy > 0.05
    ws_dash[f'F{row_num}'] = f'=IF(ABS(E{row_num})>0.05, "REVIEW", "OK")'
    
    # Apply styling
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_dash[f'{col}{row_num}'].border = border_style
        if col in ['B', 'C', 'D', 'E']:
            ws_dash[f'{col}{row_num}'].number_format = currency_format
            ws_dash[f'{col}{row_num}'].font = Font(color="000000") # Black for formulas
    
    # Input/Link color coding (State name is effectively an input/reference here)
    ws_dash[f'A{row_num}'].font = Font(color="0000FF", bold=True)
    
    row_num += 1

# Auto-size columns
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_dash.column_dimensions[col].width = 20

wb.save(output_excel)
print(f"Generated {output_excel}")
