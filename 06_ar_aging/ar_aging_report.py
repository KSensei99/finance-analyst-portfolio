import sqlite3
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
import win32com.client

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "..", "data", "nexoria.db")
    output_path = os.path.join(base_dir, "ar_aging_report.xlsx")

    # Connect to SQLite
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Query 1: Customer Summary Table
    summary_query = """
        SELECT 
            c.customer_id,
            c.company_name,
            SUM(CASE WHEN o.aging_bucket = 'Current' THEN o.invoice_total ELSE 0 END) as val_current,
            SUM(CASE WHEN o.aging_bucket = '1-30 Days' THEN o.invoice_total ELSE 0 END) as val_1_30,
            SUM(CASE WHEN o.aging_bucket = '31-60 Days' THEN o.invoice_total ELSE 0 END) as val_31_60,
            SUM(CASE WHEN o.aging_bucket = '61-90 Days' THEN o.invoice_total ELSE 0 END) as val_61_90,
            SUM(CASE WHEN o.aging_bucket = '90+ Days' THEN o.invoice_total ELSE 0 END) as val_90_plus,
            SUM(o.invoice_total) as val_total
        FROM ar_open_invoices o
        JOIN customers c ON o.customer_id = c.customer_id
        GROUP BY c.customer_id, c.company_name
        ORDER BY val_total DESC;
    """
    cur.execute(summary_query)
    summary_rows = cur.fetchall()

    # Query 2: Detailed Open Invoices Table
    detailed_query = """
        SELECT 
            o.order_id,
            o.customer_id,
            c.company_name,
            o.order_date,
            o.due_date,
            o.payment_terms,
            o.days_outstanding,
            o.aging_bucket,
            o.invoice_total
        FROM ar_open_invoices o
        JOIN customers c ON o.customer_id = c.customer_id
        ORDER BY o.days_outstanding DESC, o.order_id;
    """
    cur.execute(detailed_query)
    detailed_rows = cur.fetchall()

    # Create Workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # remove default sheet

    ws_sum = wb.create_sheet(title="AR_Summary")
    ws_det = wb.create_sheet(title="Open_Invoices_Detail")

    # Styling constants
    CLR_NAVY = "1B3A6B"
    CLR_WHITE = "FFFFFF"
    CLR_LABEL_BG = "F8FAFC"
    CLR_SECTION_BG = "EDF2F7"
    CLR_TEXT_DARK = "2D3748"
    CLR_TOTAL_BG = "F1F5F9"
    
    thin = Side(style="thin", color="CBD5E0")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_total = Border(top=Side(style="thin", color="1B3A6B"), bottom=Side(style="double", color="1B3A6B"))
    
    font_title = Font(name="Calibri", bold=True, size=15, color=CLR_WHITE)
    font_subtitle = Font(name="Calibri", italic=True, size=10, color=CLR_WHITE)
    font_header = Font(name="Calibri", bold=True, size=10, color=CLR_WHITE)
    font_data = Font(name="Calibri", size=9, color=CLR_TEXT_DARK)
    font_total = Font(name="Calibri", bold=True, size=10, color=CLR_TEXT_DARK)

    # -------------------------------------------------------------
    # SHEET 1: AR Summary Sheet
    # -------------------------------------------------------------
    ws_sum.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_sum.merge_cells("A1:H1")
    ws_sum["A1"].value = "NEXORIA COMMERCE INC."
    ws_pl = ws_sum["A1"]
    ws_pl.font = font_title
    ws_pl.fill = PatternFill("solid", fgColor=CLR_NAVY)
    ws_pl.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 26

    ws_sum.merge_cells("A2:H2")
    ws_sum["A2"].value = "Accounts Receivable (AR) Aging Summary Report"
    ws_sub = ws_sum["A2"]
    ws_sub.font = font_subtitle
    ws_sub.fill = PatternFill("solid", fgColor=CLR_NAVY)
    ws_sub.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 18

    # Executive KPI Summary Cards (Rows 4-5)
    # Total Outstanding | Overdue Balance | % Overdue
    ws_sum["A4"].value = "Total AR Outstanding"
    ws_sum["A5"].value = "=H21" # references total cell below
    ws_sum["C4"].value = "Total Overdue (>0 Days)"
    ws_sum["C5"].value = "=SUM(D21:G21)"
    ws_sum["E4"].value = "Collection Risk % (90+ Days)"
    ws_sum["E5"].value = "=G21/H21"

    # KPI Sizing and Fonts
    for col in ["A", "C", "E"]:
        ws_sum[f"{col}4"].font = Font(name="Calibri", bold=True, size=9, color="718096")
        ws_sum[f"{col}4"].alignment = Alignment(horizontal="center")
        ws_sum[f"{col}5"].font = Font(name="Calibri", bold=True, size=14, color=CLR_NAVY)
        ws_sum[f"{col}5"].alignment = Alignment(horizontal="center")
        
        # formatting KPI values
        if col in ["A", "C"]:
            ws_sum[f"{col}5"].number_format = "$#,##0"
        else:
            ws_sum[f"{col}5"].number_format = "0.0%"

    # Spacing row
    ws_sum.row_dimensions[6].height = 10

    # Column Headers (Row 7)
    ws_sum.row_dimensions[7].height = 22
    headers = [
        "Customer ID", "Customer Name", "Current ($)", "1-30 Days ($)", 
        "31-60 Days ($)", "61-90 Days ($)", "90+ Days ($)", "Total Open ($)"
    ]
    for idx, h in enumerate(headers):
        cell = ws_sum.cell(row=7, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor=CLR_NAVY)
        cell.alignment = Alignment(horizontal="center" if idx != 1 else "left", vertical="center")
        cell.border = border_all

    # Write Data rows (Rows 8 to 20 - top 13 customers)
    data_start = 8
    num_custs = min(len(summary_rows), 13)
    for idx in range(num_custs):
        r = summary_rows[idx]
        r_idx = data_start + idx
        ws_sum.row_dimensions[r_idx].height = 18
        
        # Cust ID
        c0 = ws_sum.cell(row=r_idx, column=1, value=r[0])
        c0.font = font_data
        c0.alignment = Alignment(horizontal="center")
        c0.border = border_all
        
        # Company Name
        c1 = ws_sum.cell(row=r_idx, column=2, value=r[1])
        c1.font = font_data
        c1.alignment = Alignment(horizontal="left")
        c1.border = border_all
        
        # Aging Buckets
        for bucket_idx in range(5):
            c_val = ws_sum.cell(row=r_idx, column=bucket_idx+3, value=r[bucket_idx+2])
            c_val.font = font_data
            c_val.number_format = "$#,##0;($#,##0);-"
            c_val.alignment = Alignment(horizontal="right")
            c_val.border = border_all
            
        # Total Formula
        c_tot = ws_sum.cell(row=r_idx, column=8, value=f"=SUM(C{r_idx}:G{r_idx})")
        c_tot.font = Font(name="Calibri", bold=True, size=9, color=CLR_TEXT_DARK)
        c_tot.number_format = "$#,##0;($#,##0);-"
        c_tot.alignment = Alignment(horizontal="right")
        c_tot.border = border_all

    # Total Row (Row 21)
    tot_row = data_start + num_custs
    ws_sum.row_dimensions[tot_row].height = 20
    ws_sum.cell(row=tot_row, column=1, value="").border = border_total
    
    cell_label = ws_sum.cell(row=tot_row, column=2, value="Total Accounts Receivable")
    cell_label.font = font_total
    cell_label.border = border_total
    cell_label.fill = PatternFill("solid", fgColor=CLR_TOTAL_BG)
    
    for c in range(3, 9):
        col_letter = openpyxl.utils.get_column_letter(c)
        cell_t = ws_sum.cell(row=tot_row, column=c, value=f"=SUM({col_letter}8:{col_letter}{tot_row-1})")
        cell_t.font = font_total
        cell_t.number_format = "$#,##0;($#,##0);-"
        cell_t.alignment = Alignment(horizontal="right")
        cell_t.border = border_total
        cell_t.fill = PatternFill("solid", fgColor=CLR_TOTAL_BG)

    # -------------------------------------------------------------
    # SHEET 2: Detailed Invoices Sheet
    # -------------------------------------------------------------
    ws_det.views.sheetView[0].showGridLines = True
    
    # Detailed Headers (Row 1)
    ws_det.row_dimensions[1].height = 22
    det_headers = [
        "Order ID", "Customer ID", "Company Name", "Order Date", 
        "Due Date", "Payment Terms", "Days Outstanding", "Aging Bucket", "Open Amount ($)"
    ]
    for idx, h in enumerate(det_headers):
        cell = ws_det.cell(row=1, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor=CLR_NAVY)
        cell.alignment = Alignment(horizontal="center" if idx not in [2] else "left", vertical="center")
        cell.border = border_all

    # Write Detailed Rows
    for idx, r in enumerate(detailed_rows):
        r_idx = idx + 2
        ws_det.row_dimensions[r_idx].height = 16
        for c_idx, val in enumerate(r):
            cell = ws_det.cell(row=r_idx, column=c_idx+1, value=val)
            cell.font = font_data
            cell.border = border_all
            
            # Format columns
            if c_idx in [0, 1, 3, 4, 5, 7]:
                cell.alignment = Alignment(horizontal="center")
            elif c_idx in [6]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif c_idx in [8]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "$#,##0.00;($#,##0.00);-"

    # -------------------------------------------------------------
    # Conditional Formatting Warnings (AR Summary)
    # -------------------------------------------------------------
    # Red for 90+ Days collection risk (Values > 0)
    red_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    red_font = Font(name="Calibri", bold=True, size=9, color="C5221F")
    rule_red = CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=True, fill=red_fill, font=red_font)
    
    # Orange/Yellow for 61-90 Days (Values > 0)
    orange_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    orange_font = Font(name="Calibri", bold=True, size=9, color="92400E")
    rule_orange = CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=True, fill=orange_fill, font=orange_font)
    
    # Light Green for Current bucket (Values > 0)
    green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    green_font = Font(name="Calibri", size=9, color="137333")
    rule_green = CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=True, fill=green_fill, font=green_font)

    # Apply Rules
    # Current (Col C) -> Green
    ws_sum.conditional_formatting.add(f"C8:C{tot_row-1}", rule_green)
    # 61-90 Days (Col F) -> Orange
    ws_sum.conditional_formatting.add(f"F8:F{tot_row-1}", rule_orange)
    # 90+ Days (Col G) -> Red
    ws_sum.conditional_formatting.add(f"G8:G{tot_row-1}", rule_red)

    # Save initial workbook
    wb.save(output_path)
    print(f"AR Aging report created using openpyxl: {output_path}")

    # Use COM Automation to Polish
    print("Initiating Windows COM Automation to compile formulas and auto-fit columns...")
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        xl_wb = excel.Workbooks.Open(output_path)
        excel.CalculateFull()
        
        # Auto-fit columns for both sheets
        xl_wb.Sheets("AR_Summary").Columns.AutoFit()
        xl_wb.Sheets("Open_Invoices_Detail").Columns.AutoFit()
        
        xl_wb.Save()
        xl_wb.Close()
        print("Excel COM automation completed successfully. Columns auto-fitted.")
    except Exception as e:
        print(f"COM Automation Warning/Error: {e}")
    finally:
        excel.Quit()
        conn.close()

if __name__ == "__main__":
    main()
