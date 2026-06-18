"""
=============================================================
  Nexoria Commerce Inc. — Invoice Template Builder
  Finance Analyst Portfolio | Phase 1 | Module: xlsx-official
=============================================================
  Builds a professional Excel invoice template following
  industry-standard color coding (xlsx-official skill):
    Blue  = hardcoded inputs
    Black = formulas/calculations
    Green = cross-sheet links
=============================================================
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import os

# ─── Color Palette (xlsx-official standard) ─────────────────
CLR_HEADER_BG   = "1B3A6B"   # Dark navy header
CLR_HEADER_FG   = "FFFFFF"   # White text on header
CLR_ACCENT      = "2E86AB"   # Teal accent line
CLR_INPUT_FG    = "0000FF"   # Blue = hardcoded inputs
CLR_FORMULA_FG  = "000000"   # Black = formulas
CLR_LABEL_BG    = "F0F4F8"   # Light blue-grey for labels
CLR_ALT_ROW     = "F7FAFC"   # Alternating row
CLR_TOTAL_BG    = "1B3A6B"   # Navy total row
CLR_TOTAL_FG    = "FFFFFF"
CLR_BORDER      = "B0BEC5"   # Soft border

def thin_border(color=CLR_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=CLR_BORDER):
    return Border(bottom=Side(style="thin", color=color))

def thick_bottom(color=CLR_HEADER_BG):
    return Border(bottom=Side(style="medium", color=color))

def make_font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def build_invoice_template(output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ── Column widths ────────────────────────────────────────
    col_widths = {"A": 14, "B": 30, "C": 18, "D": 12, "E": 16, "F": 16, "G": 16}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Row heights ──────────────────────────────────────────
    for row in range(1, 50):
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[1].height = 60   # Logo row
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[11].height = 20  # Column headers
    ws.row_dimensions[28].height = 20  # Totals

    # ══════════════════════════════════════════════════════════
    # SECTION 1 — HEADER (Rows 1–3)
    # ══════════════════════════════════════════════════════════
    # Set header values BEFORE merging to avoid MergedCell read-only errors
    ws["A1"].value = "NEXORIA COMMERCE INC."
    ws["A1"].font = Font(name="Calibri", bold=True, size=22, color=CLR_HEADER_FG)
    ws["A1"].fill = make_fill(CLR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A1:D1")

    ws["A2"].value = "1234 Commerce Way, Suite 100 | Austin, TX 78701"
    ws["A2"].font = make_font(size=9, color=CLR_HEADER_FG, italic=True)
    ws["A2"].fill = make_fill(CLR_HEADER_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("A2:D3")

    ws["E1"].value = "INVOICE"
    ws["E1"].font = Font(name="Calibri", bold=True, size=28, color=CLR_HEADER_FG)
    ws["E1"].fill = make_fill(CLR_HEADER_BG)
    ws["E1"].alignment = center()
    ws.merge_cells("E1:G1")

    ws["E2"].value = "www.nexoriacommerce.com"
    ws["E2"].font = Font(name="Calibri", italic=True, size=9, color=CLR_HEADER_FG)
    ws["E2"].fill = make_fill(CLR_ACCENT)
    ws["E2"].alignment = center()
    ws.merge_cells("E2:G2")

    ws["E3"].value = "billing@nexoriacommerce.com | (800) 555-0100"
    ws["E3"].font = Font(name="Calibri", italic=True, size=9, color=CLR_HEADER_FG)
    ws["E3"].fill = make_fill(CLR_ACCENT)
    ws["E3"].alignment = center()
    ws.merge_cells("E3:G3")

    # Accent divider row 4
    ws.merge_cells("A4:G4")
    ws["A4"].fill = make_fill(CLR_ACCENT)
    ws.row_dimensions[4].height = 5

    # ══════════════════════════════════════════════════════════
    # SECTION 2 — INVOICE META (Rows 5–8)
    # ══════════════════════════════════════════════════════════
    meta_left = [
        ("A5", "Invoice #:", "B5", "INV-000001"),
        ("A6", "Order Ref:", "B6", "ORD-000001"),
        ("A7", "Invoice Date:", "B7", "2024-01-15"),
        ("A8", "Due Date:", "B8", "2024-02-14"),
    ]
    for label_cell, label, val_cell, val in meta_left:
        ws[label_cell].value = label
        ws[label_cell].font = make_font(bold=True, size=9, color="374151")
        ws[label_cell].fill = make_fill(CLR_LABEL_BG)
        ws[label_cell].alignment = right()
        ws[label_cell].border = thin_border()

        ws[val_cell].value = val
        ws[val_cell].font = make_font(size=10, color=CLR_INPUT_FG)  # Blue = input
        ws[val_cell].border = thin_border()
        ws[val_cell].alignment = left()

    meta_right = [
        ("E5", "Status:", "F5", "PAID"),
        ("E6", "Terms:", "F6", "Net 30"),
        ("E7", "Carrier:", "F7", "FedEx"),
        ("E8", "Tracking #:", "F8", "ABC123456789"),
    ]
    for label_cell, label, val_cell, val in meta_right:
        ws[label_cell].value = label
        ws[label_cell].font = make_font(bold=True, size=9, color="374151")
        ws[label_cell].fill = make_fill(CLR_LABEL_BG)
        ws[label_cell].alignment = right()
        ws[label_cell].border = thin_border()

        ws[val_cell].value = val
        ws[val_cell].font = make_font(size=10, color=CLR_INPUT_FG)
        ws[val_cell].border = thin_border()
        ws[val_cell].alignment = left()

    # Merge G col for meta section
    ws.merge_cells("G5:G8")

    # ══════════════════════════════════════════════════════════
    # SECTION 3 — BILL TO / SHIP TO (Rows 9–10)
    # ══════════════════════════════════════════════════════════
    ws.row_dimensions[9].height = 14
    ws.row_dimensions[10].height = 40

    for col in ["A", "B", "C", "D"]:
        ws[f"{col}9"].fill = make_fill(CLR_LABEL_BG)
        ws[f"{col}9"].border = thin_border()
    ws.merge_cells("A9:D9")
    ws["A9"].value = "BILL TO"
    ws["A9"].font = make_font(bold=True, size=9, color=CLR_ACCENT)
    ws["A9"].alignment = left()

    ws.merge_cells("A10:D10")
    ws["A10"].value = "John Smith\nAcme Corp\n123 Main St, New York, NY 10001"
    ws["A10"].font = make_font(size=9, color=CLR_INPUT_FG)  # Blue = input
    ws["A10"].alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True, indent=1)
    ws["A10"].border = thin_border()

    for col in ["E", "F", "G"]:
        ws[f"{col}9"].fill = make_fill(CLR_LABEL_BG)
        ws[f"{col}9"].border = thin_border()
    ws.merge_cells("E9:G9")
    ws["E9"].value = "SHIP TO (if different)"
    ws["E9"].font = make_font(bold=True, size=9, color=CLR_ACCENT)
    ws["E9"].alignment = left()

    ws.merge_cells("E10:G10")
    ws["E10"].value = "Same as billing address"
    ws["E10"].font = make_font(size=9, color=CLR_INPUT_FG)
    ws["E10"].alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True, indent=1)
    ws["E10"].border = thin_border()

    # ══════════════════════════════════════════════════════════
    # SECTION 4 — LINE ITEMS TABLE (Rows 11–27)
    # ══════════════════════════════════════════════════════════
    headers = ["SKU", "Description", "Category", "Qty", "Unit Price ($)", "Line Total ($)", "Taxable?"]
    header_cols = ["A", "B", "C", "D", "E", "F", "G"]

    for col_letter, header in zip(header_cols, headers):
        cell = ws[f"{col_letter}11"]
        cell.value = header
        cell.font = Font(name="Calibri", bold=True, size=9, color=CLR_HEADER_FG)
        cell.fill = make_fill(CLR_HEADER_BG)
        cell.alignment = center()
        cell.border = thin_border()

    # 15 data rows (rows 12–26) — alternating fill
    for row in range(12, 27):
        bg = CLR_ALT_ROW if row % 2 == 0 else "FFFFFF"
        for col in header_cols:
            c = ws[f"{col}{row}"]
            c.fill = make_fill(bg)
            c.border = thin_border()
            c.alignment = center()

        # Formula for Line Total = Qty * Unit Price (Black = formula)
        ws[f"D{row}"].number_format = "#,##0"
        ws[f"E{row}"].number_format = "$#,##0.00;($#,##0.00);-"
        ws[f"F{row}"].value = f"=IFERROR(D{row}*E{row},\"\")"
        ws[f"F{row}"].font = make_font(size=9, color=CLR_FORMULA_FG)  # Black = formula
        ws[f"F{row}"].number_format = "$#,##0.00;($#,##0.00);-"

    # Sample first row data (blue = inputs)
    sample_row = 12
    ws[f"A{sample_row}"].value = "PRD-XXXX"
    ws[f"B{sample_row}"].value = "Product Description"
    ws[f"C{sample_row}"].value = "Electronics"
    ws[f"D{sample_row}"].value = 1
    ws[f"E{sample_row}"].value = 0.00
    ws[f"G{sample_row}"].value = "Yes"
    for col in ["A", "B", "C", "D", "E", "G"]:
        ws[f"{col}{sample_row}"].font = make_font(size=9, color=CLR_INPUT_FG)

    # ══════════════════════════════════════════════════════════
    # SECTION 5 — TOTALS BLOCK (Rows 28–33)
    # ══════════════════════════════════════════════════════════
    totals = [
        (28, "Subtotal:", f'=IFERROR(SUM(F12:F26),"")'),
        (29, "Taxable Amount:", '=IFERROR(SUMIF(G12:G26,"Yes",F12:F26),"")'),
        (30, "Tax Rate (%):", 0.00),
        (31, "Sales Tax:", f'=IFERROR(G29*G30,"")'),
        (32, "Shipping:", 0.00),
        (33, "INVOICE TOTAL:", f'=IFERROR(G28+G31+G32,"")'),
    ]

    for row, label, formula in totals:
        # Label cells (E:F merged)
        ws.merge_cells(f"E{row}:F{row}")
        lc = ws[f"E{row}"]
        lc.value = label
        lc.border = thin_border()
        lc.alignment = right()

        # Value cell
        vc = ws[f"G{row}"]
        is_total_row = (row == 33)
        is_input = isinstance(formula, float) and not is_total_row

        if is_total_row:
            lc.font = make_font(bold=True, size=10, color=CLR_HEADER_FG)
            lc.fill = make_fill(CLR_TOTAL_BG)
            vc.font = make_font(bold=True, size=11, color=CLR_HEADER_FG)
            vc.fill = make_fill(CLR_TOTAL_BG)
        elif is_input:
            lc.font = make_font(bold=True, size=9, color="374151")
            lc.fill = make_fill(CLR_LABEL_BG)
            vc.font = make_font(size=10, color=CLR_INPUT_FG)
            vc.fill = make_fill(CLR_LABEL_BG)
        else:
            lc.font = make_font(bold=True, size=9, color="374151")
            lc.fill = make_fill(CLR_LABEL_BG)
            vc.font = make_font(size=10, color=CLR_FORMULA_FG)  # Black = formula
            vc.fill = make_fill(CLR_LABEL_BG)

        vc.value = formula
        vc.border = thin_border()
        vc.alignment = right()

        # Number formats
        if row == 30:
            vc.number_format = "0.000%"
        else:
            vc.number_format = "$#,##0.00;($#,##0.00);-"

        # Fill blank columns A–D for total rows
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = thin_border()
            if row == 33:
                ws[f"{col}{row}"].fill = make_fill(CLR_TOTAL_BG)

    # Remap formula cells to column B (for actual generated invoices)
    # These are formula columns G → keep for template. B col holds data in generated files.

    # ══════════════════════════════════════════════════════════
    # SECTION 6 — FOOTER (Row 35–37)
    # ══════════════════════════════════════════════════════════
    ws.merge_cells("A35:G35")
    ws["A35"].fill = make_fill(CLR_ACCENT)
    ws.row_dimensions[35].height = 5

    ws.merge_cells("A36:G36")
    ws["A36"].value = "Thank you for your business! Payment via ACH, Wire, or Check. For questions: billing@nexoriacommerce.com"
    ws["A36"].font = make_font(size=8, color="6B7280", italic=True)
    ws["A36"].alignment = center()

    ws.merge_cells("A37:G37")
    ws["A37"].value = (
        "Make checks payable to: Nexoria Commerce Inc. | Bank: First National Bank | "
        "Routing: 021000021 | Account: 1234567890"
    )
    ws["A37"].font = make_font(size=8, color="6B7280")
    ws["A37"].alignment = center()

    # ── Print settings ───────────────────────────────────────
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.print_area = "A1:G37"
    ws.sheet_view.showGridLines = False

    # ── Freeze panes ─────────────────────────────────────────
    ws.freeze_panes = "A12"

    wb.save(output_path)
    print(f"Invoice template saved: {output_path}")


if __name__ == "__main__":
    out = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "invoice_template.xlsx"
    )
    build_invoice_template(out)
