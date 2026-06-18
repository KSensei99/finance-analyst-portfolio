import sqlite3
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

def copy_sheet(src_ws, dst_ws):
    # Copy merged cells
    for rng in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(rng))

    # Copy cells and styles
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font = copy(cell.font)
                dst_cell.fill = copy(cell.fill)
                dst_cell.alignment = copy(cell.alignment)
                dst_cell.border = copy(cell.border)
                dst_cell.number_format = copy(cell.number_format)

    # Copy row dimensions
    for r_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[r_idx].height = dim.height

    # Copy col dimensions
    for c_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[c_letter].width = dim.width

    # Copy print settings & view settings
    dst_ws.page_setup.orientation = src_ws.page_setup.orientation
    dst_ws.page_setup.paperSize = src_ws.page_setup.paperSize
    dst_ws.page_margins.left = src_ws.page_margins.left
    dst_ws.page_margins.right = src_ws.page_margins.right
    dst_ws.print_area = src_ws.print_area
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.freeze_panes = src_ws.freeze_panes

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "..", "data", "nexoria.db")
    template_path = os.path.join(base_dir, "invoice_template.xlsx")
    output_path = os.path.join(base_dir, "NexoriaInvoicing.xlsx")

    # Connect to SQLite
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Create new workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. Copy InvoiceTemplate sheet
    src_wb = openpyxl.load_workbook(template_path)
    src_ws = src_wb.active
    ws_template = wb.create_sheet(title="InvoiceTemplate")
    copy_sheet(src_ws, ws_template)
    src_wb.close()

    # 2. Extract & Write TaxRates
    cur.execute("SELECT state_code, state_tax_rate FROM state_tax_rates WHERE nexus = 1;")
    tax_rates = cur.fetchall()
    ws_tax = wb.create_sheet(title="TaxRates")
    ws_tax.append(["State", "Tax Rate"])
    ws_tax.column_dimensions["A"].width = 12
    ws_tax.column_dimensions["B"].width = 16
    for state, rate in tax_rates:
        cell = ws_tax.cell(row=ws_tax.max_row+1, column=1, value=state)
        cell.font = Font(name="Calibri", size=10)
        cell_rate = ws_tax.cell(row=ws_tax.max_row, column=2, value=rate)
        cell_rate.font = Font(name="Calibri", size=10)
        cell_rate.number_format = "0.000%"

    # 3. Extract & Write Orders Data (limit to 30 unique orders)
    cur.execute("""
        SELECT o.order_id 
        FROM orders o 
        WHERE o.status != 'Cancelled' 
        ORDER BY o.order_date, o.order_id 
        LIMIT 30;
    """)
    order_ids = [r[0] for r in cur.fetchall()]

    query = """
        SELECT 
            o.order_id,
            o.customer_id,
            c.company_name,
            c.state,
            p.sku,
            p.product_name,
            oi.quantity,
            oi.unit_price,
            oi.line_total,
            oi.unit_price * oi.quantity * o.tax_rate as item_tax,
            o.carrier,
            '' as invoiced,
            o.order_date,
            o.due_date,
            o.payment_terms,
            o.order_id || 'TRK' as tracking_number,
            c.address,
            c.city,
            c.zip_code,
            p.category,
            p.is_taxable,
            o.shipping_cost
        FROM order_items oi
        JOIN orders o ON oi.order_id = o.order_id
        JOIN customers c ON o.customer_id = c.customer_id
        JOIN products p ON oi.product_id = p.product_id
        WHERE o.order_id IN ({})
        ORDER BY o.order_date, o.order_id, oi.item_id;
    """.format(",".join([f"'{oid}'" for oid in order_ids]))
    
    cur.execute(query)
    rows = cur.fetchall()

    ws_orders = wb.create_sheet(title="Orders")
    headers = [
        "Order ID", "Customer ID", "Company Name", "State", "SKU", 
        "Product Name", "Quantity", "Unit Price", "Line Total", "Tax Rate", 
        "Carrier", "Invoiced", "Order Date", "Due Date", "Payment Terms", 
        "Tracking Number", "Address", "City", "Zip Code", "Category", 
        "Is Taxable", "Shipping Cost"
    ]
    ws_orders.append(headers)

    # Format header row
    header_fill = PatternFill("solid", fgColor="1B3A6B")
    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws_orders.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        row_idx = ws_orders.max_row + 1
        for c_idx, val in enumerate(r):
            cell = ws_orders.cell(row=row_idx, column=c_idx+1, value=val)
            cell.font = Font(name="Calibri", size=9)
            
            # Formatting specifics
            if c_idx in [0, 1, 3, 4, 10, 11, 12, 13, 14, 15, 18]:
                cell.alignment = Alignment(horizontal="center")
            elif c_idx in [6]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif c_idx in [7, 8, 21]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "$#,##0.00;($#,##0.00);-"
            elif c_idx in [9]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.000%"

    # Auto-adjust column widths
    for col in ws_orders.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_orders.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # 4. Create Log Sheet
    ws_log = wb.create_sheet(title="Log")
    log_headers = ["Invoice #", "Company Name", "Timestamp", "PDF Path"]
    ws_log.append(log_headers)
    for col_idx in range(1, len(log_headers) + 1):
        cell = ws_log.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    ws_log.column_dimensions["A"].width = 16
    ws_log.column_dimensions["B"].width = 30
    ws_log.column_dimensions["C"].width = 20
    ws_log.column_dimensions["D"].width = 60

    # Save workbook
    wb.save(output_path)
    print(f"Invoicing automation workbook saved: {output_path}")
    conn.close()

if __name__ == "__main__":
    main()
