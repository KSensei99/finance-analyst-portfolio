import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import win32com.client

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    actuals_path = os.path.join(base_dir, "actuals.csv")
    budget_path = os.path.join(base_dir, "budget.csv")
    output_path = os.path.join(base_dir, "Nexoria_PL_Report.xlsx")

    # 1. Create openpyxl Workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Add sheets
    ws_pl = wb.create_sheet(title="PL_Report")
    ws_act = wb.create_sheet(title="Actuals")
    ws_bud = wb.create_sheet(title="Budget")
    ws_set = wb.create_sheet(title="Settings")

    # Load and write raw Actuals data
    with open(actuals_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            # Convert year, month, and amount to numbers
            if row[0] == "year":
                ws_act.append(row)
            else:
                ws_act.append([int(row[0]), int(row[1]), row[2], row[3], row[4], float(row[5])])

    # Load and write raw Budget data
    with open(budget_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0] == "year":
                ws_bud.append(row)
            else:
                ws_bud.append([int(row[0]), int(row[1]), row[2], row[3], row[4], float(row[5])])

    # Generate unique periods list for dropdown
    periods = []
    for y in [2023, 2024]:
        for m in range(1, 13):
            periods.append(f"{y}-{m:02d}")

    # Write Settings
    ws_set.append(["Period"])
    for p in periods:
        ws_set.append([p])

    # 2. Design PL_Report Layout
    ws_pl.views.sheetView[0].showGridLines = True

    # Color Palette
    CLR_NAVY = "1B3A6B"
    CLR_WHITE = "FFFFFF"
    CLR_LABEL_BG = "F7FAFC"
    CLR_SECTION_BG = "EDF2F7"
    CLR_ACCENT_GOLD = "C5A880"
    CLR_LIGHT_GOLD = "F7F5F0"
    CLR_TEXT_DARK = "2D3748"
    CLR_FORMULA = "000000"
    CLR_INPUT = "0000FF"

    # Borders
    thin = Side(style="thin", color="D2D6DC")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_top_bottom = Border(top=thin, bottom=thin)
    border_total = Border(top=Side(style="thin", color="1B3A6B"), bottom=Side(style="double", color="1B3A6B"))
    border_gold = Border(bottom=Side(style="medium", color=CLR_ACCENT_GOLD))

    # Fonts
    font_title = Font(name="Calibri", bold=True, size=16, color=CLR_WHITE)
    font_subtitle = Font(name="Calibri", italic=True, size=11, color=CLR_WHITE)
    font_section = Font(name="Calibri", bold=True, size=11, color=CLR_TEXT_DARK)
    font_total = Font(name="Calibri", bold=True, size=10, color=CLR_TEXT_DARK)
    font_ebit = Font(name="Calibri", bold=True, size=11, color=CLR_WHITE)
    font_data = Font(name="Calibri", size=10, color=CLR_TEXT_DARK)
    font_formula = Font(name="Calibri", size=10, color=CLR_FORMULA)
    font_input = Font(name="Calibri", bold=True, size=10, color=CLR_INPUT)

    # Column dimensions
    ws_pl.column_dimensions["A"].width = 32
    ws_pl.column_dimensions["B"].width = 16
    ws_pl.column_dimensions["C"].width = 16
    ws_pl.column_dimensions["D"].width = 16
    ws_pl.column_dimensions["E"].width = 16

    # 3. Header Title Block
    ws_pl.merge_cells("A1:E1")
    ws_pl["A1"].value = "NEXORIA COMMERCE INC."
    ws_pl["A1"].font = font_title
    ws_pl["A1"].fill = PatternFill("solid", fgColor=CLR_NAVY)
    ws_pl["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_pl.row_dimensions[1].height = 28

    ws_pl.merge_cells("A2:E2")
    ws_pl["A2"].value = "Monthly Profit & Loss Statement (Actual vs. Budget)"
    ws_pl["A2"].font = font_subtitle
    ws_pl["A2"].fill = PatternFill("solid", fgColor=CLR_NAVY)
    ws_pl["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_pl.row_dimensions[2].height = 20

    # 4. Slicer Month Selector
    ws_pl["A3"].value = "Select Period (YYYY-MM):"
    ws_pl["A3"].font = Font(name="Calibri", bold=True, size=10, color=CLR_TEXT_DARK)
    ws_pl["A3"].alignment = Alignment(horizontal="right", vertical="center")

    ws_pl["B3"].value = "2024-12" # Default value (Blue text = input)
    ws_pl["B3"].font = font_input
    ws_pl["B3"].alignment = Alignment(horizontal="center", vertical="center")
    ws_pl["B3"].fill = PatternFill("solid", fgColor=CLR_LIGHT_GOLD)
    ws_pl["B3"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Helper cells for formulas (D3 Year, E3 Month)
    ws_pl["D3"].value = "=VALUE(LEFT(B3,4))"
    ws_pl["D3"].font = Font(color="FFFFFF") # hide text by making it white
    ws_pl["E3"].value = "=VALUE(RIGHT(B3,2))"
    ws_pl["E3"].font = Font(color="FFFFFF") # hide text

    # Data Validation Dropdown for B3
    dv = DataValidation(type="list", formula1="Settings!$A$2:$A$25", allow_blank=True)
    dv.error = 'Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select a period from the list'
    dv.promptTitle = 'Select Period'
    ws_pl.add_data_validation(dv)
    dv.add(ws_pl["B3"])

    # 5. Table Headers
    ws_pl.row_dimensions[5].height = 22
    headers = ["Line Item", "Actual ($)", "Budget ($)", "Variance ($)", "Variance (%)"]
    for idx, h in enumerate(headers):
        cell = ws_pl.cell(row=5, column=idx+1, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=CLR_WHITE)
        cell.fill = PatternFill("solid", fgColor=CLR_NAVY)
        cell.alignment = Alignment(horizontal="center" if idx > 0 else "left", vertical="center")
        cell.border = border_all

    # Format mapping helpers
    c_fmt = "$#,##0;($#,##0);-"
    p_fmt = "0.0%"

    # 6. Build Row Structure
    # Map row index to details: (Type: "section"|"item"|"total", Label, account_name/formula_range, val_sign)
    row_struct = {
        7: ("section", "REVENUE", "", 1),
        8: ("item", "  Product Sales", "Product Sales", 1),
        9: ("item", "  Service Revenue", "Service Revenue", 1),
        10: ("item", "  Shipping Revenue", "Shipping Revenue", 1),
        11: ("total", "Total Revenue", "8:10", 1),
        
        13: ("section", "COST OF GOODS SOLD", "", -1),
        14: ("item", "  Product Cost", "Product Cost", -1),
        15: ("item", "  Freight In", "Freight In", -1),
        16: ("item", "  Packaging", "Packaging", -1),
        17: ("total", "Total Cost of Goods Sold", "14:16", -1),
        
        19: ("section_total", "GROSS PROFIT", "11-17", 1),
        
        21: ("section", "OPERATING EXPENSES", "", -1),
        22: ("item", "  Salaries", "Salaries", -1),
        23: ("item", "  Marketing Spend", "Marketing Spend", -1),
        24: ("item", "  Software & SaaS", "Software & SaaS", -1),
        25: ("item", "  Office Rent", "Office Rent", -1),
        26: ("item", "  Travel & Entertainment", "Travel & Entertainment", -1),
        27: ("item", "  Professional Services", "Professional Services", -1),
        28: ("item", "  Utilities", "Utilities", -1),
        29: ("item", "  Depreciation", "Depreciation", -1),
        30: ("total", "Total Operating Expenses", "22:29", -1),
        
        32: ("ebit", "NET OPERATING INCOME (EBIT)", "19-30", 1),
    }

    for r_idx, (r_type, label, val, sign) in row_struct.items():
        ws_pl.row_dimensions[r_idx].height = 20
        cell_a = ws_pl.cell(row=r_idx, column=1, value=label)
        
        if r_type == "section":
            cell_a.font = font_section
            cell_a.fill = PatternFill("solid", fgColor=CLR_SECTION_BG)
            cell_a.alignment = Alignment(horizontal="left", vertical="center")
            # Fill other columns with empty background
            for c in range(2, 6):
                ws_pl.cell(row=r_idx, column=c).fill = PatternFill("solid", fgColor=CLR_SECTION_BG)
                ws_pl.cell(row=r_idx, column=c).border = border_top_bottom
                
        elif r_type == "item":
            cell_a.font = font_data
            cell_a.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            
            # Actual Formula (Col B)
            cell_b = ws_pl.cell(row=r_idx, column=2, value=f'=SUMIFS(Actuals!F:F, Actuals!A:A, $D$3, Actuals!B:B, $E$3, Actuals!E:E, "{val}")')
            cell_b.font = font_formula
            cell_b.number_format = c_fmt
            cell_b.alignment = Alignment(horizontal="right", vertical="center")
            
            # Budget Formula (Col C)
            cell_c = ws_pl.cell(row=r_idx, column=3, value=f'=SUMIFS(Budget!F:F, Budget!A:A, $D$3, Budget!B:B, $E$3, Budget!E:E, "{val}")')
            cell_c.font = font_formula
            cell_c.number_format = c_fmt
            cell_c.alignment = Alignment(horizontal="right", vertical="center")
            
            # Variance $ Formula (Col D)
            # Favorable = Actual - Budget for Revenue (sign=1)
            # Favorable = Budget - Actual for Expenses (sign=-1)
            if sign == 1:
                cell_d = ws_pl.cell(row=r_idx, column=4, value=f"=B{r_idx}-C{r_idx}")
            else:
                cell_d = ws_pl.cell(row=r_idx, column=4, value=f"=C{r_idx}-B{r_idx}")
            cell_d.font = font_formula
            cell_d.number_format = c_fmt
            cell_d.alignment = Alignment(horizontal="right", vertical="center")
            
            # Variance % Formula (Col E)
            cell_e = ws_pl.cell(row=r_idx, column=5, value=f"=IFERROR(D{r_idx}/C{r_idx}, 0)")
            cell_e.font = font_formula
            cell_e.number_format = p_fmt
            cell_e.alignment = Alignment(horizontal="right", vertical="center")
            
        elif r_type == "total":
            cell_a.font = font_total
            cell_a.alignment = Alignment(horizontal="left", vertical="center")
            
            start_row, end_row = val.split(":")
            # Actual Sum
            cell_b = ws_pl.cell(row=r_idx, column=2, value=f"=SUM(B{start_row}:B{end_row})")
            cell_b.font = font_total
            cell_b.number_format = c_fmt
            cell_b.alignment = Alignment(horizontal="right", vertical="center")
            cell_b.border = border_total
            
            # Budget Sum
            cell_c = ws_pl.cell(row=r_idx, column=3, value=f"=SUM(C{start_row}:C{end_row})")
            cell_c.font = font_total
            cell_c.number_format = c_fmt
            cell_c.alignment = Alignment(horizontal="right", vertical="center")
            cell_c.border = border_total
            
            # Variance $
            if sign == 1:
                cell_d = ws_pl.cell(row=r_idx, column=4, value=f"=B{r_idx}-C{r_idx}")
            else:
                cell_d = ws_pl.cell(row=r_idx, column=4, value=f"=C{r_idx}-B{r_idx}")
            cell_d.font = font_total
            cell_d.number_format = c_fmt
            cell_d.alignment = Alignment(horizontal="right", vertical="center")
            cell_d.border = border_total
            
            # Variance %
            cell_e = ws_pl.cell(row=r_idx, column=5, value=f"=IFERROR(D{r_idx}/C{r_idx}, 0)")
            cell_e.font = font_total
            cell_e.number_format = p_fmt
            cell_e.alignment = Alignment(horizontal="right", vertical="center")
            cell_e.border = border_total
            
        elif r_type == "section_total":
            # Gross Profit = Total Revenue (row 11) - Total COGS (row 17)
            cell_a.font = font_total
            cell_a.alignment = Alignment(horizontal="left", vertical="center")
            cell_a.fill = PatternFill("solid", fgColor=CLR_LIGHT_GOLD)
            cell_a.border = border_gold
            
            cell_b = ws_pl.cell(row=r_idx, column=2, value="=B11-B17")
            cell_b.font = font_total
            cell_b.number_format = c_fmt
            cell_b.alignment = Alignment(horizontal="right", vertical="center")
            cell_b.fill = PatternFill("solid", fgColor=CLR_LIGHT_GOLD)
            cell_b.border = border_gold
            
            cell_c = ws_pl.cell(row=r_idx, column=3, value="=C11-C17")
            cell_c.font = font_total
            cell_c.number_format = c_fmt
            cell_c.alignment = Alignment(horizontal="right", vertical="center")
            cell_c.fill = PatternFill("solid", fgColor=CLR_LIGHT_GOLD)
            cell_c.border = border_gold
            
            cell_d = ws_pl.cell(row=r_idx, column=4, value="=B19-C19")
            cell_d.font = font_total
            cell_d.number_format = c_fmt
            cell_d.alignment = Alignment(horizontal="right", vertical="center")
            cell_d.fill = PatternFill("solid", fgColor=CLR_LIGHT_GOLD)
            cell_d.border = border_gold
            
            cell_e = ws_pl.cell(row=r_idx, column=5, value="=IFERROR(D19/C19, 0)")
            cell_e.font = font_total
            cell_e.number_format = p_fmt
            cell_e.alignment = Alignment(horizontal="right", vertical="center")
            cell_e.fill = PatternFill("solid", fgColor=CLR_LIGHT_GOLD)
            cell_e.border = border_gold
            
        elif r_type == "ebit":
            # Operating Income = Gross Profit (row 19) - Total OpEx (row 30)
            ws_pl.row_dimensions[r_idx].height = 24
            cell_a.font = font_ebit
            cell_a.alignment = Alignment(horizontal="left", vertical="center")
            cell_a.fill = PatternFill("solid", fgColor=CLR_NAVY)
            
            cell_b = ws_pl.cell(row=r_idx, column=2, value="=B19-B30")
            cell_b.font = font_ebit
            cell_b.number_format = c_fmt
            cell_b.alignment = Alignment(horizontal="right", vertical="center")
            cell_b.fill = PatternFill("solid", fgColor=CLR_NAVY)
            
            cell_c = ws_pl.cell(row=r_idx, column=3, value="=C19-C30")
            cell_c.font = font_ebit
            cell_c.number_format = c_fmt
            cell_c.alignment = Alignment(horizontal="right", vertical="center")
            cell_c.fill = PatternFill("solid", fgColor=CLR_NAVY)
            
            cell_d = ws_pl.cell(row=r_idx, column=4, value="=B32-C32")
            cell_d.font = font_ebit
            cell_d.number_format = c_fmt
            cell_d.alignment = Alignment(horizontal="right", vertical="center")
            cell_d.fill = PatternFill("solid", fgColor=CLR_NAVY)
            
            cell_e = ws_pl.cell(row=r_idx, column=5, value="=IFERROR(D32/C32, 0)")
            cell_e.font = font_ebit
            cell_e.number_format = p_fmt
            cell_e.alignment = Alignment(horizontal="right", vertical="center")
            cell_e.fill = PatternFill("solid", fgColor=CLR_NAVY)

    # 7. Add Conditional Formatting for Variance columns D and E
    # Light Green fill: E6F4EA, Text: 137333
    # Light Red fill: FCE8E6, Text: C5221F
    fill_green = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    font_green = Font(name="Calibri", bold=True, size=10, color="137333")
    fill_red = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    font_red = Font(name="Calibri", bold=True, size=10, color="C5221F")

    # Greater than 0 is Favorable
    rule_green = CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=fill_green, font=font_green)
    # Less than 0 is Unfavorable
    rule_red = CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=fill_red, font=font_red)

    # Apply to item and total variance columns (D and E)
    # We apply them row-range wise to avoid coloring blank or ebit navy background cells
    ws_pl.conditional_formatting.add("D8:E12", rule_green)
    ws_pl.conditional_formatting.add("D8:E12", rule_red)
    ws_pl.conditional_formatting.add("D14:E18", rule_green)
    ws_pl.conditional_formatting.add("D14:E18", rule_red)
    ws_pl.conditional_formatting.add("D19:E19", rule_green)
    ws_pl.conditional_formatting.add("D19:E19", rule_red)
    ws_pl.conditional_formatting.add("D22:E31", rule_green)
    ws_pl.conditional_formatting.add("D22:E31", rule_red)

    # Save initial workbook
    wb.save(output_path)
    print(f"Excel workbook created using openpyxl: {output_path}")

    # 8. Use Windows COM Automation to polish and compile
    print("Initiating Windows COM Automation to compile formulas and auto-fit columns...")
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        xl_wb = excel.Workbooks.Open(output_path)
        # Recalculate all formulas in Excel context
        excel.CalculateFull()
        # Auto-fit columns in the report sheet
        xl_ws = xl_wb.Sheets("PL_Report")
        xl_ws.Columns.AutoFit()
        
        # Save and close workbook
        xl_wb.Save()
        xl_wb.Close()
        print("Excel COM automation completed successfully. Calculations compiled, column widths auto-fit.")
    except Exception as e:
        print(f"COM Automation Warning/Error: {e}")
    finally:
        excel.Quit()

if __name__ == "__main__":
    main()
