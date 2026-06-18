import sqlite3
import pandas as pd
from openpyxl.chart import PieChart, Reference
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

db_path = r'C:\Users\marsh\Downloads\Upwork\finance-analyst-portfolio\data\nexoria.db'
output_dir = r'C:\Users\marsh\Downloads\Upwork\finance-analyst-portfolio\03_shipping_allocation'

# 1. Read Data
conn = sqlite3.connect(db_path)
shipments = pd.read_sql_query("SELECT * FROM shipments", conn)
orders = pd.read_sql_query("SELECT order_id, total_weight_lbs FROM orders", conn)
conn.close()

# 2. Allocate Shipping Costs
allocation_data = []

for _, shipment in shipments.iterrows():
    s_id = shipment['shipment_id']
    carrier = shipment['carrier']
    billed_cost = shipment['billed_shipping_cost']
    
    # Split bundled orders
    order_ids = [oid.strip() for oid in str(shipment['order_id']).split(',')]
    
    if len(order_ids) == 1:
        # 100% allocation
        allocation_data.append({
            'shipment_id': s_id,
            'order_id': order_ids[0],
            'carrier': carrier,
            'allocated_cost': billed_cost,
            'allocation_pct': 1.0
        })
    else:
        # Proportional allocation by weight
        bundle_orders = orders[orders['order_id'].isin(order_ids)]
        total_bundle_weight = bundle_orders['total_weight_lbs'].sum()
        
        for _, order in bundle_orders.iterrows():
            weight = order['total_weight_lbs']
            if total_bundle_weight > 0:
                alloc_pct = weight / total_bundle_weight
            else:
                alloc_pct = 1.0 / len(order_ids) # Fallback if weight is 0
                
            allocated_cost = billed_cost * alloc_pct
            
            allocation_data.append({
                'shipment_id': s_id,
                'order_id': order['order_id'],
                'carrier': carrier,
                'allocated_cost': round(allocated_cost, 2),
                'allocation_pct': round(alloc_pct, 4)
            })

alloc_df = pd.DataFrame(allocation_data)

# Save Raw CSV
csv_path = f"{output_dir}\\shipping_allocation_report.csv"
alloc_df.to_csv(csv_path, index=False)

# 3. Create Excel Summary
excel_path = f"{output_dir}\\shipping_allocation_summary.xlsx"
wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Carrier Summary"
ws_data = wb.create_sheet(title="Allocation Data")

# Write Data
for r in dataframe_to_rows(alloc_df, index=False, header=True):
    ws_data.append(r)

# Aggregate Summary by Carrier
summary_df = alloc_df.groupby('carrier').agg(
    total_cost=('allocated_cost', 'sum'),
    shipment_count=('shipment_id', 'nunique'),
    order_count=('order_id', 'count')
).reset_index()

summary_df['avg_cost_per_order'] = summary_df['total_cost'] / summary_df['order_count']
summary_df = summary_df.round(2)

ws_summary.append(["Carrier", "Total Shipping Cost", "Unique Shipments", "Orders Serviced", "Avg Cost per Order"])
for r in dataframe_to_rows(summary_df, index=False, header=False):
    ws_summary.append(r)

# Formatting Excel
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True)
currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

for cell in ws_summary[1]:
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center")

for row in range(2, ws_summary.max_row + 1):
    ws_summary[f'B{row}'].number_format = currency_format
    ws_summary[f'E{row}'].number_format = currency_format

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_summary.column_dimensions[col].width = 20

# 4. Create Native Excel Pie Chart
pie = PieChart()
labels = Reference(ws_summary, min_col=1, min_row=2, max_row=ws_summary.max_row)
data = Reference(ws_summary, min_col=2, min_row=1, max_row=ws_summary.max_row)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Total Shipping Cost Breakdown by Carrier"

# Add it to the summary sheet at column G
ws_summary.add_chart(pie, "G2")

wb.save(excel_path)

print(f"Generated {csv_path}")
print(f"Generated {excel_path} (with native Pie Chart)")
